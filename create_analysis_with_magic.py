"""
Create Analysis_{date}.xlsx with:
1. Fixed ATR handling (weekend-aware)
2. Magic number breakdown columns in RESULT tab
3. #code mapping columns for easy reference
"""
import MetaTrader5 as mt5
import pandas as pd
from datetime import datetime, timedelta, date, timezone
from dotenv import load_dotenv
import os
from collections import defaultdict
import sys
sys.path.append('C:/Trading_GU')
from magic_code_config import MAGIC_CODES, get_code_info

# PT multipliers per magic number (for correct ATRTP calculation)
PT_MULTIPLIERS = {m: info['pt_mult'] for m, info in MAGIC_CODES.items()}

load_dotenv()

def connect_mt5_terminal(terminal_key):
    """Connect to MT5 terminal using env var"""
    path = os.getenv(terminal_key)
    if not path or not mt5.initialize(path=path):
        return False
    return True

def get_atr_m1_60(target_time):
    """Get ATR(M1,60) - use available history if < 60 candles"""
    symbol = 'XAUUSDp'
    timeframe = mt5.TIMEFRAME_M1
    
    # Check if target_time is Monday (weekday() == 0)
    # XAUUSD doesn't trade on Sundays, so look back to Friday
    target_weekday = target_time.weekday()
    if target_weekday == 0:  # Monday
        hours_back = 72  # Look back to Friday
    else:
        hours_back = 3
    
    # Keep timezone-aware for API
    utc_from = target_time - timedelta(hours=hours_back)
    utc_to = target_time + timedelta(minutes=5)
    
    rates = mt5.copy_rates_range(symbol, timeframe, utc_from, utc_to)
    if rates is None or len(rates) < 10:  # Need at least 10 candles
        return None
    
    df = pd.DataFrame(rates)
    df['time'] = pd.to_datetime(df['time'], unit='s')
    
    # Filter to candles before target time
    target_time_naive = target_time.replace(tzinfo=None) if target_time.tzinfo else target_time
    df = df[df['time'] <= target_time_naive]
    
    if len(df) < 10:
        return None
    
    # Calculate ATR with available candles
    df['high_low'] = df['high'] - df['low']
    df['high_close'] = abs(df['high'] - df['close'].shift())
    df['low_close'] = abs(df['low'] - df['close'].shift())
    df['tr'] = df[['high_low', 'high_close', 'low_close']].max(axis=1)
    
    # Use available periods for ATR (min 14, max 60)
    atr_periods = min(60, len(df) - 1)
    if atr_periods < 14:
        atr_periods = 14
    
    atr = df['tr'].rolling(atr_periods).mean().iloc[-1]
    return atr if not pd.isna(atr) else None

def fetch_gu_positions(target_date):
    """Fetch GU positions from Vantage"""
    date_from = datetime.combine(target_date, datetime.min.time(), tzinfo=timezone.utc)
    date_to = date_from + timedelta(days=1)
    
    deals = mt5.history_deals_get(date_from, date_to)
    if not deals:
        return []
    
    positions = []
    deals_by_pos = defaultdict(list)
    for d in deals:
        deals_by_pos[d.position_id].append(d)
    
    for pid, siblings in deals_by_pos.items():
        entry_deal, exit_deal = None, None
        for s in siblings:
            if s.entry == 0:
                entry_deal = s
            elif s.entry == 1:
                exit_deal = s
        
        if entry_deal and exit_deal:
            is_gu = ('GU_' in str(entry_deal.comment)) or str(entry_deal.magic).startswith('282603')
            if is_gu:
                direction = "BUY" if entry_deal.type == 0 else "SELL"
                positions.append({
                    'pos_id': pid,
                    'magic': entry_deal.magic,
                    'direction': direction,
                    'open_time': datetime.utcfromtimestamp(entry_deal.time).replace(tzinfo=timezone.utc),
                    'close_time': datetime.utcfromtimestamp(exit_deal.time).replace(tzinfo=timezone.utc),
                    'entry_price': entry_deal.price,
                    'close_price': exit_deal.price,
                    'volume': entry_deal.volume
                })
    
    return positions

def calculate_mae_mfe(tick_data, entry_price, direction, minutes, open_time):
    """Calculate MFE/MAE for a time window"""
    # Make naive for comparison
    open_time_naive = open_time.replace(tzinfo=None) if open_time.tzinfo else open_time
    window_end = open_time_naive + timedelta(minutes=minutes)
    
    mask = (tick_data['time'] >= open_time_naive) & (tick_data['time'] <= window_end)
    window_ticks = tick_data[mask]
    
    if window_ticks.empty:
        return None, 0, None, 0, None
    
    if direction == 'BUY':
        mfe_price = window_ticks['ask'].max()
        mfe_points = int(round((mfe_price - entry_price) * 100))
        mae_price = window_ticks['bid'].min()
        mae_points = int(round((entry_price - mae_price) * 100))
        close_price = window_ticks['bid'].iloc[-1]
    else:
        mfe_price = window_ticks['bid'].min()
        mfe_points = int(round((entry_price - mfe_price) * 100))
        mae_price = window_ticks['ask'].max()
        mae_points = int(round((mae_price - entry_price) * 100))
        close_price = window_ticks['ask'].iloc[-1]
    
    return mfe_price, mfe_points, mae_price, mae_points, close_price

def main(target_date=None):
    if target_date is None:
        target_date = date(2026, 3, 24)  # Default to March 24th
    
    print("=" * 80)
    print(f"Creating Analysis_{target_date.strftime('%Y%m%d')}.xlsx with Magic Number Breakdown")
    print("=" * 80)
    
    # Step 1: Fetch positions from Vantage
    print("\n1. Connecting to Vantage...")
    if not connect_mt5_terminal('MT5_TERMINAL_VANTAGE'):
        print("Failed")
        return
    
    print("2. Fetching positions...")
    positions = fetch_gu_positions(target_date)
    print(f"   Found {len(positions)} GU positions")
    
    # Get unique magic numbers
    magic_numbers = sorted(set(p['magic'] for p in positions))
    print(f"   Magic numbers found: {magic_numbers}")
    
    mt5.shutdown()
    
    # Step 2: Connect to BlackBull for tick data and ATR
    print("\n3. Connecting to BlackBull...")
    if not connect_mt5_terminal('MT5_TERMINAL_BLACKBULL'):
        print("Failed")
        return
    
    print("4. Analyzing positions...")
    all_results = {}  # time_window -> list of position results
    magic_outcomes = {m: {} for m in magic_numbers}  # magic -> time_window -> total_outcome
    
    for m in magic_numbers:
        for min_val in range(1, 31):
            magic_outcomes[m][f"{min_val}min"] = 0
    
    for i, pos in enumerate(positions):
        # Get ATR
        atr_value = get_atr_m1_60(pos['open_time'])
        if atr_value is None:
            atr_value = 3.5  # Default ATR if calculation fails
        
        # Calculate ATRTP using correct PT multiplier for this magic number
        pt_mult = PT_MULTIPLIERS.get(pos['magic'], 0.5)
        atr_tp_points = int(round(atr_value * pt_mult * 100))
        
        # Get tick data (30 min window)
        open_time = pos['open_time']
        window_end = open_time + timedelta(minutes=30)
        ticks = mt5.copy_ticks_range('XAUUSDp', open_time, window_end, mt5.COPY_TICKS_ALL)
        
        if ticks is None or len(ticks) == 0:
            continue
        
        df_ticks = pd.DataFrame(ticks)
        df_ticks['time'] = pd.to_datetime(df_ticks['time'], unit='s')
        
        # Calculate actual points
        if pos['direction'] == 'BUY':
            actual_points = int(round((pos['close_price'] - pos['entry_price']) * 100))
        else:
            actual_points = int(round((pos['entry_price'] - pos['close_price']) * 100))
        
        # Analyze each time window
        for minutes in range(1, 31):
            mfe_price, mfe_points, mae_price, mae_points, close_at_window = calculate_mae_mfe(
                df_ticks, pos['entry_price'], pos['direction'], minutes, open_time
            )
            
            if mfe_price is None:
                continue
            
            # Determine outcome
            if mfe_points >= atr_tp_points:
                outcome = "PROFIT"
                outcome_points = atr_tp_points
            else:
                outcome = "LOSS"
                if pos['direction'] == 'BUY':
                    outcome_points = int(round((close_at_window - pos['entry_price']) * 100))
                else:
                    outcome_points = int(round((pos['entry_price'] - close_at_window) * 100))
            
            # Store result
            time_key = f"{minutes}min"
            if time_key not in all_results:
                all_results[time_key] = []
            
            all_results[time_key].append({
                'Ticket': f"P{pos['pos_id']}",
                'Magic Number': pos['magic'],
                'Type': pos['direction'],
                'TimeOpen': pos['open_time'].strftime('%Y-%m-%d %H:%M:%S'),
                'PriceOpen': pos['entry_price'],
                'TimeClose': pos['close_time'].strftime('%Y-%m-%d %H:%M:%S'),
                'PriceClose': pos['close_price'],
                'ActualPoints': actual_points,
                'ATROpen': round(atr_value, 2),
                'ATRTP': atr_tp_points,
                f'MFE{minutes}Price': mfe_price,
                f'MFE{minutes}Points': mfe_points,
                f'MAE{minutes}Price': mae_price,
                f'MAE{minutes}Points': mae_points,
                'Outcome': outcome,
                'OutcomePoints': outcome_points
            })
            
            # Accumulate by magic number
            magic_outcomes[pos['magic']][time_key] += outcome_points
        
        if (i + 1) % 50 == 0:
            print(f"   Processed {i+1}/{len(positions)} positions...")
    
    mt5.shutdown()
    
    # Step 3: Create Excel
    print("\n5. Creating Excel output...")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Create RESULT sheet with magic number columns
    ws_result = wb.active
    ws_result.title = "RESULT"
    
    # Headers: TimeWindow + standard cols + magic number cols
    base_headers = ['TimeWindow', 'ProfitCount', 'LossCount', 'WinRate', 'TotalOutcomePoints']
    magic_headers = [f'Magic{m}' for m in magic_numbers]
    headers = base_headers + magic_headers
    
    for col, header in enumerate(headers, 1):
        cell = ws_result.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Fill RESULT data
    for row_idx, minutes in enumerate(range(1, 31), 2):
        time_key = f"{minutes}min"
        
        if time_key in all_results:
            results = all_results[time_key]
            profits = sum(1 for r in results if r['Outcome'] == 'PROFIT')
            losses = sum(1 for r in results if r['Outcome'] == 'LOSS')
            total = profits + losses
            win_rate = (profits / total * 100) if total > 0 else 0
            total_outcome = sum(r['OutcomePoints'] for r in results)
            
            ws_result.cell(row=row_idx, column=1, value=time_key)
            ws_result.cell(row=row_idx, column=2, value=profits)
            ws_result.cell(row=row_idx, column=3, value=losses)
            ws_result.cell(row=row_idx, column=4, value=round(win_rate, 2))
            ws_result.cell(row=row_idx, column=5, value=total_outcome)
            
            # Add magic number outcomes
            for col_idx, magic in enumerate(magic_numbers, 6):
                ws_result.cell(row=row_idx, column=col_idx, value=magic_outcomes[magic][time_key])
    
    # Create per-minute sheets
    for minutes in range(1, 31):
        time_key = f"{minutes}min"
        
        if time_key not in all_results:
            continue
        
        ws = wb.create_sheet(title=time_key)
        results = all_results[time_key]
        
        if not results:
            continue
        
        # Get columns from first result and add #Code column
        cols = ['#Code'] + list(results[0].keys())
        
        # Headers
        for col, header in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Data
        for row_idx, result in enumerate(results, 2):
            # Add #Code as first column
            magic = result['Magic Number']
            code_info = get_code_info(magic)
            ws.cell(row=row_idx, column=1, value=code_info['code'])
            
            # Add rest of data
            for col, key in enumerate(cols[1:], 2):  # Skip #Code in results, start at column 2
                ws.cell(row=row_idx, column=col, value=result[key])
    
    # Save (use _v4 suffix to avoid file lock issues)
    output_file = f"data/Analysis_{target_date.strftime('%Y%m%d')}_v4.xlsx"
    wb.save(output_file)
    
    print(f"\nExcel saved: {output_file}")
    print(f"\nSummary:")
    print(f"  - Total positions: {len(positions)}")
    print(f"  - Magic numbers: {magic_numbers}")
    print(f"  - Sheets: RESULT + 30 time windows")
    print(f"  - RESULT columns: {headers}")
    print(f"  - Per-minute columns: #Code + standard columns")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Create Analysis Excel for a specific date')
    parser.add_argument('--date', type=str, help='Date in YYYY-MM-DD format (default: 2026-03-24)')
    args = parser.parse_args()
    
    if args.date:
        from datetime import datetime
        target_date = datetime.strptime(args.date, '%Y-%m-%d').date()
        main(target_date)
    else:
        main()
