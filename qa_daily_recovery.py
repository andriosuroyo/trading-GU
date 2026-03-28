"""
Daily RecoveryAnalysis Generator (CommentTag-Based)
Generates {date}_RecoveryAnalysis.xlsx for RGU validation

Usage: python qa_daily_recovery.py --date 2026-03-27
Output: data/20260327_RecoveryAnalysis.xlsx
"""
import MetaTrader5 as mt5
import pandas as pd
import numpy as np
from datetime import datetime, timedelta, timezone
from collections import defaultdict
import os
import argparse

# Configuration
OUTPUT_DIR = "data"
TRADING_WINDOW_START = 2
TRADING_WINDOW_END = 23
BASKET_WINDOW_SECONDS = 300  # 5-minute window
MAX_LAYERS = 5

# Recovery analysis matrix: time windows (hours) × ATR multipliers
RECOVERY_HOURS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24]  # 2H to 24H
ATR_MULTIPLIERS = [1.0, 2.0, 3.0]  # 1x (aggressive), 2x (balanced), 3x (conservative)


def load_env():
    """Load environment variables from .env file."""
    env_vars = {}
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    env_vars[key.strip()] = val.strip()
    return env_vars


def connect_mt5_terminal(terminal_key, max_retries=3):
    """Connect to MT5 terminal with retry logic."""
    import time
    env_vars = load_env()
    terminal_path = env_vars.get(terminal_key)
    
    if not terminal_path or not os.path.exists(terminal_path):
        print(f"Error: Terminal path not found for {terminal_key}")
        return False
    
    for attempt in range(max_retries):
        # Ensure any previous connection is fully closed
        try:
            mt5.shutdown()
            time.sleep(1)  # Give terminal time to release
        except:
            pass
        
        if mt5.initialize(path=terminal_path):
            info = mt5.account_info()
            print(f"Connected to: {info.server} | Account: {info.login}")
            return True
        
        error = mt5.last_error()
        print(f"MT5 initialize() failed (attempt {attempt + 1}/{max_retries}): {error}")
        
        if attempt < max_retries - 1:
            time.sleep(2)  # Wait before retry
    
    return False


def is_gu_position(comment):
    """Check if position is GU strategy based on CommentTag."""
    if comment and "GU_" in str(comment).upper():
        return True
    return False


def get_session_from_time(dt):
    """Derive trading session from OpenTime (UTC)."""
    hour = dt.hour
    if 2 <= hour < 6:
        return "Asia"
    elif 8 <= hour < 12:
        return "London"
    elif 17 <= hour < 21:
        return "NY"
    else:
        return "Other"


def is_carry_over_trade(open_time, close_time):
    """Check if position is carry-over."""
    close_hour = close_time.hour
    return close_hour < TRADING_WINDOW_START or close_hour > TRADING_WINDOW_END


def fetch_gu_positions(target_date):
    """Fetch all GU positions from MT5 for target date using CommentTag."""
    date_from = datetime.combine(target_date, datetime.min.time(), tzinfo=timezone.utc)
    date_to = date_from + timedelta(days=1)
    
    print(f"\nFetching positions for {target_date}...")
    
    deals = mt5.history_deals_get(date_from, date_to)
    if not deals:
        print("No deals found")
        return []
    
    print(f"Total deals: {len(deals)}")
    
    deals_by_pos = defaultdict(list)
    for d in deals:
        deals_by_pos[d.position_id].append(d)
    
    positions = []
    
    for pid, siblings in deals_by_pos.items():
        entry_deal = None
        exit_deal = None
        
        for s in siblings:
            if s.entry == 0:
                entry_deal = s
            elif s.entry == 1:
                exit_deal = s
        
        if entry_deal and exit_deal:
            if not is_gu_position(entry_deal.comment):
                continue
            
            direction = "BUY" if entry_deal.type == 0 else "SELL"
            open_time = datetime.utcfromtimestamp(entry_deal.time).replace(tzinfo=timezone.utc)
            close_time = datetime.utcfromtimestamp(exit_deal.time).replace(tzinfo=timezone.utc)
            
            session = get_session_from_time(open_time)
            profit = exit_deal.profit + entry_deal.commission + entry_deal.commission + exit_deal.swap
            
            positions.append({
                "ticket": pid,
                "magic": entry_deal.magic,
                "session": session,
                "direction": direction,
                "open_time": open_time,
                "close_time": close_time,
                "open_price": entry_deal.price,
                "close_price": exit_deal.price,
                "volume": entry_deal.volume,
                "profit": profit,
                "comment": entry_deal.comment
            })
    
    print(f"GU positions found: {len(positions)}")
    return positions


def identify_loss_baskets(positions):
    """
    Group losing positions into baskets (within 5-minute window, same direction).
    Returns list of baskets.
    """
    losing = [p for p in positions if p["profit"] < 0]
    losing.sort(key=lambda x: x["open_time"])
    
    baskets = []
    
    if not losing:
        return baskets
    
    current_basket = [losing[0]]
    basket_start_time = losing[0]["open_time"]
    basket_direction = losing[0]["direction"]
    
    for i in range(1, len(losing)):
        pos = losing[i]
        time_diff = (pos["open_time"] - basket_start_time).total_seconds()
        
        if pos["direction"] == basket_direction and time_diff <= BASKET_WINDOW_SECONDS:
            current_basket.append(pos)
        else:
            baskets.append(current_basket)
            current_basket = [pos]
            basket_start_time = pos["open_time"]
            basket_direction = pos["direction"]
    
    if current_basket:
        baskets.append(current_basket)
    
    return baskets


def get_atr_m1_60(target_time):
    """Get ATR(60) on M1 timeframe at target time from BlackBull."""
    symbol_bb = "XAUUSDp"
    from_time = target_time - timedelta(hours=2)
    rates = mt5.copy_rates_range(symbol_bb, mt5.TIMEFRAME_M1, from_time, target_time)
    
    if rates is None or len(rates) < 60:
        return None
    
    df = pd.DataFrame(rates)
    df["time"] = pd.to_datetime(df["time"], unit="s", utc=True)
    
    df["high_low"] = df["high"] - df["low"]
    df["high_close"] = abs(df["high"] - df["close"].shift())
    df["low_close"] = abs(df["low"] - df["close"].shift())
    df["tr"] = df[["high_low", "high_close", "low_close"]].max(axis=1)
    df["atr_60"] = df["tr"].rolling(window=60).mean()
    
    df_before = df[df["time"] <= target_time]
    if df_before.empty:
        return None
    
    return df_before["atr_60"].iloc[-1]


def fetch_m1_candles_for_window(start_time, end_time):
    """Fetch M1 candle data from BlackBull for time window."""
    symbol_bb = "XAUUSDp"
    # Add small buffer to ensure we get all candles
    fetch_start = start_time - timedelta(minutes=1)
    fetch_end = end_time + timedelta(minutes=1)
    
    rates = mt5.copy_rates_range(symbol_bb, mt5.TIMEFRAME_M1, fetch_start, fetch_end)
    if rates is None or len(rates) == 0:
        return None
    
    df = pd.DataFrame(rates)
    df["time"] = pd.to_datetime(df["time"], unit="s", utc=True)
    return df


def get_price_from_candles(candle_df, target_time, direction, price_type="close"):
    """Get price from M1 candles at specific time.
    
    Args:
        candle_df: DataFrame with M1 candle data
        target_time: Target datetime
        direction: "BUY" or "SELL" (for choosing bid/ask if needed)
        price_type: "open", "high", "low", or "close"
    """
    if candle_df is None:
        return None
    
    # Find candle at or before target_time
    mask = candle_df["time"] <= target_time
    if not mask.any():
        return None
    
    closest_candle = candle_df[mask].iloc[-1]
    
    if price_type == "open":
        return closest_candle["open"]
    elif price_type == "high":
        return closest_candle["high"]
    elif price_type == "low":
        return closest_candle["low"]
    else:  # close
        return closest_candle["close"]


def analyze_basket_for_window(basket, all_positions, basket_num, recovery_minutes, atr_multiplier=2.0):
    """
    Analyze a loss basket for a specific recovery window.
    
    Args:
        basket: List of positions in the basket
        all_positions: All positions from the day
        basket_num: Basket number (1-based)
        recovery_minutes: Recovery window in minutes
        atr_multiplier: ATR multiplier for layer spacing (1.0=aggressive, 2.0=balanced, 3.0=conservative)
    """
    if not basket:
        return None
    
    primary = basket[0]
    
    # Format basket_id: B001, B002, etc.
    basket_id = f"B{basket_num:03d}"
    
    # NumPos: Count of positions in basket
    num_pos = len(basket)
    
    # Magic: List all magic numbers separated by comma
    magics = sorted(set(str(p["magic"]) for p in basket))
    magic_str = ",".join(magics)
    
    # Direction
    direction = primary["direction"]
    
    # Times
    open_time = primary["open_time"]
    close_time = primary["close_time"]
    
    # Prices
    open_price = primary["open_price"]
    close_price = primary["close_price"]
    
    # Get ATR for basket
    atr_value = get_atr_m1_60(open_time)
    if atr_value is not None and not pd.isna(atr_value):
        atr_points = atr_value * 100
    else:
        atr_points = 250  # Default
    
    entry_distance = atr_points * atr_multiplier
    
    # Recovery window
    recovery_deadline = close_time + timedelta(minutes=recovery_minutes)
    
    # Fetch M1 candle data for this window (much faster than tick data)
    candle_df = fetch_m1_candles_for_window(close_time, recovery_deadline)
    
    # Calculate Loss for this window:
    # If basket positions closed within window, use actual close price
    # Otherwise, use candle close price at window end
    window_end_price = None
    if candle_df is not None:
        window_end_price = get_price_from_candles(candle_df, recovery_deadline, direction, "close")
    
    # Calculate total loss for window
    total_loss = 0
    for pos in basket:
        if pos["close_time"] <= recovery_deadline:
            # Position closed within window - use actual loss
            total_loss += abs(pos["profit"]) * 100
        else:
            # Position still open at window end - estimate loss from candle data
            if window_end_price:
                if direction == "BUY":
                    loss = (pos["open_price"] - window_end_price) * 100
                else:
                    loss = (window_end_price - pos["open_price"]) * 100
                total_loss += max(0, loss)  # Only count if it's a loss
            else:
                # No candle data - use actual loss as fallback
                total_loss += abs(pos["profit"]) * 100
    
    # Check recovery within this window using M1 candle highs/lows
    recovered = False
    recovery_time = None
    recovery_time_dt = None
    
    if candle_df is not None:
        window_candles = candle_df[(candle_df["time"] >= close_time) & (candle_df["time"] <= recovery_deadline)]
        if not window_candles.empty:
            if direction == "BUY":
                # For BUY: recovery when price goes up to open_price (check if high reaches it)
                recovery_candles = window_candles[window_candles["high"] >= open_price]
                if not recovery_candles.empty:
                    recovered = True
                    recovery_time_dt = recovery_candles.iloc[0]["time"]
                    recovery_time = (recovery_time_dt - close_time).total_seconds() / 60
            else:  # SELL
                # For SELL: recovery when price goes down to open_price (check if low reaches it)
                recovery_candles = window_candles[window_candles["low"] <= open_price]
                if not recovery_candles.empty:
                    recovered = True
                    recovery_time_dt = recovery_candles.iloc[0]["time"]
                    recovery_time = (recovery_time_dt - close_time).total_seconds() / 60
    
    # Calculate furthest price during window using candle lows/highs
    furthest_price = None
    max_mae_points = 0
    if candle_df is not None:
        window_candles = candle_df[(candle_df["time"] >= close_time) & (candle_df["time"] <= recovery_deadline)]
        if not window_candles.empty:
            if direction == "BUY":
                # For BUY: furthest is the lowest low
                furthest_price = window_candles["low"].min()
                max_mae_points = (open_price - furthest_price) * 100 if furthest_price else 0
            else:
                # For SELL: furthest is the highest high
                furthest_price = window_candles["high"].max()
                max_mae_points = (furthest_price - open_price) * 100 if furthest_price else 0
    
    # Count opposing positions BEFORE recovery or deadline
    opposing_direction = "SELL" if direction == "BUY" else "BUY"
    opp_count_end_time = recovery_time_dt if recovery_time_dt else recovery_deadline
    
    opp_positions = [p for p in all_positions 
                    if p["direction"] == opposing_direction 
                    and close_time < p["open_time"] < opp_count_end_time]
    opp_count = len(opp_positions)
    
    # Calculate RecoveryClose price
    # If recovered: use OpenPrice (the recovery level)
    # If not recovered: use candle close price at window end
    recovery_close_price = None
    if recovered:
        recovery_close_price = open_price
    else:
        # Not recovered - get candle close price at window end
        if candle_df is not None:
            recovery_close_price = get_price_from_candles(candle_df, recovery_deadline, direction, "close")
    
    # Calculate layers - positions opened AFTER basket closes during recovery window
    # Layer 1: First same-direction position opened after close_time
    # Layer 2+: Subsequent positions with ATR spacing from previous layer
    # All layers target the original OpenPrice and close at window end if not recovered
    
    same_direction_positions = []
    for pos in all_positions:
        if pos["ticket"] == primary["ticket"]:
            continue
        if pos["direction"] != direction:
            continue
        # Must open AFTER basket closes and within recovery window
        opp_count_end_time = recovery_time_dt if recovery_time_dt else recovery_deadline
        if not (close_time < pos["open_time"] <= opp_count_end_time):
            continue
        same_direction_positions.append(pos)
    
    same_direction_positions.sort(key=lambda x: x["open_time"])
    
    layers = []
    entry_distance_price = entry_distance / 100
    last_entry_price = None
    
    for pos in same_direction_positions:
        if len(layers) == 0:
            layers.append({"price": pos["open_price"], "time": pos["open_time"]})
            last_entry_price = pos["open_price"]
        else:
            # Check ATR distance from previous layer
            distance_from_last = abs(pos["open_price"] - last_entry_price)
            
            # Check if this layer is FURTHER from OpenPrice than the last layer
            # For BUY: want LOWER prices (further down from OpenPrice)
            # For SELL: want HIGHER prices (further up from OpenPrice)
            if direction == "BUY":
                # New layer must be LOWER than last (closer to furthest price)
                is_further_from_target = pos["open_price"] < last_entry_price
            else:  # SELL
                # New layer must be HIGHER than last (closer to furthest price)
                is_further_from_target = pos["open_price"] > last_entry_price
            
            if distance_from_last >= entry_distance_price and is_further_from_target and len(layers) < MAX_LAYERS:
                layers.append({"price": pos["open_price"], "time": pos["open_time"]})
                last_entry_price = pos["open_price"]
    
    # Calculate layer data and total PL
    layer_data = {}
    total_pl = 0  # Total P&L of all layers
    
    for i, layer in enumerate(layers, 1):
        # Calculate PL based on RecoveryClose price
        if direction == "BUY":
            # BUY: profit if close > entry
            if recovery_close_price:
                pl = (recovery_close_price - layer["price"]) * 100
            else:
                pl = 0
        else:  # SELL
            # SELL: profit if close < entry
            if recovery_close_price:
                pl = (layer["price"] - recovery_close_price) * 100
            else:
                pl = 0
        
        layer_data[f"Layer{i}Price"] = round(layer["price"], 2)
        layer_data[f"Layer{i}Time"] = layer["time"].strftime("%H:%M:%S")
        layer_data[f"Layer{i}PL"] = round(pl, 1)
        
        # Add to total PL
        total_pl += pl
    
    # Fill empty layers
    for i in range(len(layers) + 1, MAX_LAYERS + 1):
        layer_data[f"Layer{i}Price"] = None
        layer_data[f"Layer{i}Time"] = None
        layer_data[f"Layer{i}PL"] = None
    

    
    # Build result row
    result = {
        "BasketID": basket_id,
        "NumPos": num_pos,
        "Magic": magic_str,
        "Direction": direction,
        "OpenTime": open_time.strftime("%H:%M:%S"),
        "CloseTime": close_time.strftime("%H:%M:%S"),
        "OpenPrice": round(open_price, 2),
        "ClosePrice": round(close_price, 2),
        "ATRPoints": round(atr_points),
        "EntryDistance": round(entry_distance),
        "OppCount": opp_count,
        "LayerCount": len(layers),
        "Recovered": "YES" if recovered else "NO",
        "RecoveryTime": round(recovery_time, 1) if recovery_time else recovery_minutes,  # Show window max if not recovered
        "RecoveryClose": round(recovery_close_price, 2) if recovery_close_price else None,
        "FurthestPrice": round(furthest_price, 2) if furthest_price else None,
        "TotalPL": round(total_pl, 2),  # Total P&L of all layers at RecoveryClose
        "MaxMAE": round(max_mae_points, 2) if max_mae_points > 0 else 0
    }
    
    # Add layer data
    result.update(layer_data)
    
    return result


def generate_by_magic_number(basket_results_matrix, all_baskets, all_positions):
    """Generate By Magic Number summary using 12H 3x as reference.
    
    IMPORTANT: When a basket contains multiple magic numbers, the basket's
    TotalPL and Recovery status are applied to ALL magic numbers in that basket.
    """
    magic_data = defaultdict(lambda: {
        "baskets": set(),
        "recovered": 0,
        "total_pl": 0
    })
    
    # Use 12H 3x as reference for By Magic Number
    ref_key = "12H 3x"
    if ref_key not in basket_results_matrix:
        return pd.DataFrame()
    
    df_ref = basket_results_matrix[ref_key]
    
    # For each basket, distribute TotalPL and Recovery status to ALL magic numbers
    for _, row in df_ref.iterrows():
        basket_id = row["BasketID"]
        magics = str(row["Magic"]).split(",")
        recovered = row["Recovered"] == "YES"
        basket_total_pl = row["TotalPL"] if "TotalPL" in row else 0
        
        for magic_str in magics:
            magic = int(magic_str.strip())
            magic_data[magic]["baskets"].add(basket_id)
            if recovered:
                magic_data[magic]["recovered"] += 1
            # Add the basket's TotalPL to each magic number's total
            magic_data[magic]["total_pl"] += basket_total_pl
    
    # Also calculate the actual trading NetPL from positions (for reference)
    for magic in magic_data:
        magic_positions = [p for p in all_positions if p["magic"] == magic]
        trading_pl = sum(p["profit"] for p in magic_positions) * 100
        magic_data[magic]["trading_pl"] = round(trading_pl, 2)
    
    rows = []
    for magic in sorted(magic_data.keys()):
        data = magic_data[magic]
        total_baskets = len(data["baskets"])
        recovery_rate = (data["recovered"] / total_baskets * 100) if total_baskets > 0 else 0
        basket_list = ",".join(sorted(data["baskets"]))
        
        rows.append({
            "Magic": magic,
            "Baskets": basket_list,
            "TotalBaskets": total_baskets,
            "Recovered": data["recovered"],
            "RecoveryRate": round(recovery_rate, 1),
            "RecoveryPL": round(data.get("total_pl", 0), 2),  # TotalPL from recovery analysis
            "TradingPL": data.get("trading_pl", 0)  # Actual position P&L for reference
        })
    
    return pd.DataFrame(rows)


def generate_recovery_analysis(target_date):
    """Main function to generate RecoveryAnalysis."""
    print("=" * 80)
    print(f"Daily RecoveryAnalysis Generator")
    print(f"Date: {target_date}")
    print("=" * 80)
    
    # Connect to Vantage
    if not connect_mt5_terminal("MT5_TERMINAL_VANTAGE"):
        return None
    
    # Fetch positions
    positions = fetch_gu_positions(target_date)
    if not positions:
        mt5.shutdown()
        return None
    
    # Filter carry-over trades
    positions = [p for p in positions if not is_carry_over_trade(p["open_time"], p["close_time"])]
    print(f"After carry-over filter: {len(positions)}")
    
    mt5.shutdown()
    import time
    time.sleep(2)  # Allow terminal to fully release before next connection
    
    # Connect to BlackBull for ATR/tick data
    if not connect_mt5_terminal("MT5_TERMINAL_BLACKBULL"):
        return None
    
    # Identify loss baskets
    baskets = identify_loss_baskets(positions)
    print(f"Loss baskets identified: {len(baskets)}")
    
    if not baskets:
        mt5.shutdown()
        print("No loss baskets to analyze")
        return None
    
    # Analyze baskets for each recovery window and ATR multiplier combination
    # Matrix: 2H-24H (2H increments) × 1x/2x/3x ATR multipliers
    basket_results_matrix = {}  # Key: "2H 1x", Value: DataFrame
    
    for hours in RECOVERY_HOURS:
        window_minutes = hours * 60
        for atr_mult in ATR_MULTIPLIERS:
            sheet_key = f"{hours}H {int(atr_mult)}x"
            print(f"\nAnalyzing {sheet_key}...")
            
            results = []
            for i, basket in enumerate(baskets, 1):
                result = analyze_basket_for_window(basket, positions, i, window_minutes, atr_mult)
                if result:
                    results.append(result)
            
            # Create DataFrame
            df = pd.DataFrame(results)
            
            # Reorder columns
            base_cols = ["BasketID", "NumPos", "Magic", "Direction", "OpenTime", "CloseTime", 
                         "OpenPrice", "ClosePrice", "ATRPoints", "EntryDistance",
                         "OppCount", "LayerCount", "Recovered", "RecoveryTime", "RecoveryClose", 
                         "FurthestPrice", "TotalPL", "MaxMAE"]
            
            layer_cols = []
            for i in range(1, MAX_LAYERS + 1):
                layer_cols.extend([f"Layer{i}Price", f"Layer{i}Time", f"Layer{i}PL"])
            
            all_cols = base_cols + layer_cols
            df = df[[col for col in all_cols if col in df.columns]]
            
            basket_results_matrix[sheet_key] = df
            
            recovered_count = sum(1 for r in results if r["Recovered"] == "YES")
            print(f"  {sheet_key}: {len(results)} baskets, {recovered_count} recovered ({recovered_count/len(results)*100:.1f}%)")
    
    mt5.shutdown()
    
    # Generate By Magic Number summary (using 12H 3x as reference)
    df_by_magic = generate_by_magic_number(basket_results_matrix, baskets, positions)
    
    # Save to Excel
    output_file = os.path.join(OUTPUT_DIR, f"{target_date.strftime('%Y%m%d')}_RecoveryAnalysis.xlsx")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Build Summary sheet with matrix data
        summary_rows = []
        
        # Create column headers for matrix
        matrix_keys = [f"{h}H {int(m)}x" for h in RECOVERY_HOURS for m in ATR_MULTIPLIERS]
        
        # Row 1: Total Baskets
        row = {"Metric": "Total Baskets"}
        for key in matrix_keys:
            row[key] = len(basket_results_matrix[key])
        summary_rows.append(row)
        
        # Row 2: Recovered Baskets
        row = {"Metric": "Recovered Baskets"}
        for key in matrix_keys:
            df = basket_results_matrix[key]
            recovered = sum(1 for _, r in df.iterrows() if r["Recovered"] == "YES")
            row[key] = recovered
        summary_rows.append(row)
        
        # Row 3: Recovery Rate (%)
        row = {"Metric": "Recovery Rate (%)"}
        for key in matrix_keys:
            df = basket_results_matrix[key]
            total = len(df)
            recovered = sum(1 for _, r in df.iterrows() if r["Recovered"] == "YES")
            rate = (recovered / total * 100) if total > 0 else 0
            row[key] = round(rate, 1)
        summary_rows.append(row)
        
        # Row 4: Total PL (sum of all layer P&L)
        row = {"Metric": "Total PL"}
        for key in matrix_keys:
            df = basket_results_matrix[key]
            total_pl = df["TotalPL"].sum() if "TotalPL" in df.columns else 0
            row[key] = round(total_pl, 2)
        summary_rows.append(row)
        
        # Row 5: Recovered PL (sum of layer P&L for recovered baskets)
        row = {"Metric": "Recovered PL"}
        for key in matrix_keys:
            df = basket_results_matrix[key]
            recovered_df = df[df["Recovered"] == "YES"]
            recovered_pl = recovered_df["TotalPL"].sum() if "TotalPL" in df.columns else 0
            row[key] = round(recovered_pl, 2)
        summary_rows.append(row)
        
        # Row 6: Lost PL (sum of layer P&L for NOT recovered baskets)
        row = {"Metric": "Lost PL"}
        for key in matrix_keys:
            df = basket_results_matrix[key]
            not_recovered_df = df[df["Recovered"] == "NO"]
            lost_pl = not_recovered_df["TotalPL"].sum() if "TotalPL" in df.columns else 0
            row[key] = round(lost_pl, 2)
        summary_rows.append(row)
        
        # Row 7: Max MAE (max across all baskets)
        row = {"Metric": "Max MAE"}
        for key in matrix_keys:
            df = basket_results_matrix[key]
            max_mae = df["MaxMAE"].max() if "MaxMAE" in df.columns else 0
            row[key] = round(max_mae, 2)
        summary_rows.append(row)
        
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Store 12H 3x data for final output reference
        df_ref = basket_results_matrix.get("12H 3x")
        total_baskets = len(df_ref) if df_ref is not None else 0
        recovered_count = sum(1 for _, r in df_ref.iterrows() if r["Recovered"] == "YES") if df_ref is not None else 0
        recovery_rate = (recovered_count / total_baskets * 100) if total_baskets > 0 else 0
        
        # Individual matrix sheets
        for key in matrix_keys:
            basket_results_matrix[key].to_excel(writer, sheet_name=key, index=False)
        
        # By Magic Number sheet
        if not df_by_magic.empty:
            df_by_magic.to_excel(writer, sheet_name='By Magic Number', index=False)
        
        # Formatting
        workbook = writer.book
        
        # Color TotalPL row in Summary sheet - highlight best (highest) in bright green
        if 'Summary' in writer.sheets:
            summary_sheet = writer.sheets['Summary']
            
            # Find Total PL row and get all values
            total_pl_row_idx = None
            total_pl_values = []
            
            for row_idx, row in enumerate(df_summary.itertuples(), 2):
                if row.Metric == "Total PL":
                    total_pl_row_idx = row_idx
                    # Get all numeric values from the row (skip Metric column)
                    for col_idx in range(2, len(df_summary.columns) + 1):
                        cell = summary_sheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            total_pl_values.append((col_idx, cell.value))
                    break
            
            # Color TotalPL cells - green for positive, red for negative
            if total_pl_values:
                from openpyxl.styles import PatternFill
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")   # Light red
                
                for col_idx, value in total_pl_values:
                    cell = summary_sheet.cell(row=total_pl_row_idx, column=col_idx)
                    if value > 0:
                        cell.fill = green_fill
                    elif value < 0:
                        cell.fill = red_fill
        
        # Color Recovered column and LayerXPL columns in each matrix sheet
        green_fill = None
        red_fill = None
        
        for key in matrix_keys:
            if key in writer.sheets:
                sheet = writer.sheets[key]
                df = basket_results_matrix[key]
                
                for col_idx, col_name in enumerate(df.columns, 1):
                    from openpyxl.styles import PatternFill
                    if green_fill is None:
                        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    
                    # Color Recovered column
                    if col_name == "Recovered":
                        for row_idx, row in enumerate(df.itertuples(), 2):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            if cell.value == "YES":
                                cell.fill = green_fill
                            elif cell.value == "NO":
                                cell.fill = red_fill
                    
                    # Color LayerXPL columns (green for positive, red for negative)
                    if col_name.startswith("Layer") and col_name.endswith("PL"):
                        for row_idx, row in enumerate(df.itertuples(), 2):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            if cell.value and isinstance(cell.value, (int, float)):
                                if cell.value > 0:
                                    cell.fill = green_fill
                                elif cell.value < 0:
                                    cell.fill = red_fill
        
        # Color RecoveryPL column in By Magic Number sheet
        if 'By Magic Number' in writer.sheets and not df_by_magic.empty:
            sheet = writer.sheets['By Magic Number']
            
            for col_idx, col_name in enumerate(df_by_magic.columns, 1):
                if col_name == "RecoveryPL":
                    
                    for row_idx, row in enumerate(df_by_magic.itertuples(), 2):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value and cell.value > 0:
                            cell.fill = green_fill
                        elif cell.value and cell.value < 0:
                            cell.fill = red_fill
                    break
    
    print("\n" + "=" * 80)
    print("RecoveryAnalysis Complete")
    print("=" * 80)
    print(f"Output: {output_file}")
    print(f"\nSheets generated:")
    print(f"  - Summary")
    for key in matrix_keys:
        print(f"  - {key}")
    print(f"  - By Magic Number")
    print(f"\nTotal baskets: {total_baskets}")
    print(f"Recovery rate (12H 3x): {recovery_rate:.1f}%")
    
    # Escalation check
    if recovery_rate < 70:
        print(f"\n[ESCALATION] Recovery rate {recovery_rate:.1f}% below 70% threshold")
    
    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Generate Daily RecoveryAnalysis')
    parser.add_argument('--date', type=str, help='Date in YYYY-MM-DD format (default: yesterday)')
    args = parser.parse_args()
    
    if args.date:
        target_date = datetime.strptime(args.date, '%Y-%m-%d').date()
    else:
        target_date = (datetime.now(timezone.utc) - timedelta(days=1)).date()
    
    generate_recovery_analysis(target_date)
