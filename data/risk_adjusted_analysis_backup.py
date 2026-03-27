"""
Risk-Adjusted Analysis: Total Outcome / Max MAE Ratio
Adds risk dimension to the comprehensive time × multiplier matrix
"""
import pandas as pd
import sys
sys.path.append('C:/Trading_GU')
from datetime import datetime, timedelta, date
from fetch_all_gu_positions import fetch_all_positions, connect_mt5
from datetime import timezone
import MetaTrader5 as mt5

# Configuration matrix
time_windows = [5, 10, 15, 20, 25, 30]
multipliers = [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]

def connect_blackbull():
    """Connect to BlackBull terminal"""
    import os
    from dotenv import load_dotenv
    load_dotenv()
    blackbull_path = os.getenv('MT5_TERMINAL_BLACKBULL')
    
    if not mt5.initialize(path=blackbull_path):
        print(f"Failed to initialize MT5: {mt5.last_error()}")
        return False
    print(f"Connected to BlackBull MT5")
    return True

def get_atr_value(target_time):
    """Get ATR(M1,60) value at specific time"""
    symbol = 'XAUUSDp'
    timeframe = mt5.TIMEFRAME_M1
    
    # Handle timezone - make naive for MT5
    if target_time.tzinfo is not None:
        target_time = target_time.replace(tzinfo=None)
    
    # Get rates around target time (need 60+ candles for ATR)
    utc_from = target_time - timedelta(hours=3)
    utc_to = target_time + timedelta(minutes=5)
    
    rates = mt5.copy_rates_range(symbol, timeframe, utc_from, utc_to)
    if rates is None or len(rates) < 60:
        return None
    
    df = pd.DataFrame(rates)
    df['time'] = pd.to_datetime(df['time'], unit='s')
    
    # Filter to candles before target time
    df = df[df['time'] <= target_time]
    if len(df) < 60:
        return None
    
    # Calculate ATR(60)
    df['high_low'] = df['high'] - df['low']
    df['high_close'] = abs(df['high'] - df['close'].shift())
    df['low_close'] = abs(df['low'] - df['close'].shift())
    df['tr'] = df[['high_low', 'high_close', 'low_close']].max(axis=1)
    atr = df['tr'].rolling(60).mean().iloc[-1]
    
    return atr

def calculate_mae_mfe_for_window(position, tick_data, minutes=15):
    """
    Calculate MFE and MAE for a specific time window
    Returns: (mfe_price, mfe_points, mae_price, mae_points)
    """
    open_time = position['open_time']
    entry_price = position['entry_price']
    direction = position['direction']
    
    # Handle timezone - make both naive for comparison
    if open_time.tzinfo is not None:
        open_time = open_time.replace(tzinfo=None)
    
    window_end = open_time + timedelta(minutes=minutes)
    
    # Filter ticks to window
    mask = (tick_data['time'] >= open_time) & (tick_data['time'] <= window_end)
    window_ticks = tick_data[mask]
    
    if window_ticks.empty:
        return None, 0, None, 0
    
    if direction == 'BUY':
        # MFE: max ask (best price to sell)
        mfe_price = window_ticks['ask'].max()
        mfe_points = int(round((mfe_price - entry_price) * 100))
        
        # MAE: min bid (worst price against position)
        mae_price = window_ticks['bid'].min()
        mae_points = int(round((entry_price - mae_price) * 100))
    else:  # SELL
        # MFE: min bid (best price to buy back)
        mfe_price = window_ticks['bid'].min()
        mfe_points = int(round((entry_price - mfe_price) * 100))
        
        # MAE: max ask (worst price against position)
        mae_price = window_ticks['ask'].max()
        mae_points = int(round((mae_price - entry_price) * 100))
    
    return mfe_price, mfe_points, mae_price, mae_points

def main(target_date=date(2026, 3, 20)):
    print("=" * 100)
    print("RISK-ADJUSTED ANALYSIS: Total Outcome / Max MAE Ratio")
    print("=" * 100)
    print(f"\nAnalyzing date: {target_date}")
    
    # Connect to Vantage and fetch positions
    print("\nConnecting to Vantage...")
    if not connect_mt5("MT5_TERMINAL_VANTAGE"):
        print("Failed to connect")
        sys.exit(1)
    
    print("\nFetching positions from Vantage...")
    positions_all = fetch_all_positions(
        datetime.combine(target_date, datetime.min.time(), tzinfo=timezone.utc),
        datetime.combine(target_date + timedelta(days=1), datetime.min.time(), tzinfo=timezone.utc)
    )
# Filter for GU positions and map field names
positions = []
for p in positions_all:
    if p['is_gu']:
        positions.append({
            'position_id': p['pos_id'],
            'magic': p['magic'],
            'direction': p['direction'],
            'open_time': p['open_time'],
            'close_time': p['close_time'],
            'entry_price': p['open_price'],
            'close_price': p['close_price'],
            'net_pl': p['net_pl'],
            'volume': p['volume']
        })
print(f"Loaded {len(positions)} GU positions")

# Connect to BlackBull for tick data
print("\nConnecting to BlackBull for tick data...")
if not connect_blackbull():
    print("Failed to connect to BlackBull")
    sys.exit(1)

# Cache tick data for each position
print("\nFetching tick data for all positions (up to 30min window)...")
tick_cache = {}
for i, pos in enumerate(positions):
    symbol = 'XAUUSDp'
    open_time = pos['open_time']
    window_end = open_time + timedelta(minutes=30)
    
    ticks = mt5.copy_ticks_range(symbol, open_time, window_end, mt5.COPY_TICKS_ALL)
    if ticks is not None and len(ticks) > 0:
        df_ticks = pd.DataFrame(ticks)
        df_ticks['time'] = pd.to_datetime(df_ticks['time'], unit='s')
        tick_cache[pos['position_id']] = df_ticks
        if (i+1) % 10 == 0:
            print(f"  Processed {i+1}/{len(positions)} positions...")
    else:
        print(f"  Position {i+1}: No tick data!")

print(f"\nCached tick data for {len(tick_cache)} positions")

# Store results
results = []

print("\n" + "=" * 100)
print("ANALYZING 54 CONFIGURATIONS (calculating MAE for each)...")
print("=" * 100)

for tw in time_windows:
    for mult in multipliers:
        config_name = f"{tw}min_{mult}x"
        
        total_outcome = 0
        profit_count = 0
        loss_count = 0
        all_mae_values = []
        
        for pos in positions:
            pos_id = pos['position_id']
            if pos_id not in tick_cache:
                continue
                
            tick_data = tick_cache[pos_id]
            
            # Get ATR at position open
            atr_value = get_atr_value(pos['open_time'])
            if atr_value is None:
                continue
                
            atr_tp_points = int(round(atr_value * mult * 100))
            
            # Calculate MFE/MAE for this time window
            mfe_price, mfe_points, mae_price, mae_points = calculate_mae_mfe_for_window(
                pos, tick_data, minutes=tw
            )
            
            if mfe_price is None:
                continue
            
            # Get close price at window end
            open_time_naive = pos['open_time'].replace(tzinfo=None) if pos['open_time'].tzinfo else pos['open_time']
            window_end = open_time_naive + timedelta(minutes=tw)
            window_ticks = tick_data[(tick_data['time'] >= open_time_naive) & (tick_data['time'] <= window_end)]
            
            if len(window_ticks) == 0:
                continue
                
            if pos['direction'] == 'BUY':
                close_price = window_ticks['bid'].iloc[-1]
            else:
                close_price = window_ticks['ask'].iloc[-1]
            
            # Determine outcome
            if mfe_points >= atr_tp_points:
                outcome = atr_tp_points
                profit_count += 1
            else:
                if pos['direction'] == 'BUY':
                    outcome_points = int(round((close_price - pos['entry_price']) * 100))
                else:
                    outcome_points = int(round((pos['entry_price'] - close_price) * 100))
                outcome = outcome_points
                loss_count += 1
            
            total_outcome += outcome
            all_mae_values.append(mae_points)
        
        # Calculate risk metrics
        if len(all_mae_values) > 0:
            max_mae = max(all_mae_values)
            avg_mae = sum(all_mae_values) / len(all_mae_values)
            
            # Risk-Adjusted Return Ratio = Total Outcome / Max MAE
            if max_mae > 0:
                risk_adjusted_ratio = total_outcome / max_mae
            else:
                risk_adjusted_ratio = 0
            
            total_trades = profit_count + loss_count
            win_rate = (profit_count / total_trades * 100) if total_trades > 0 else 0
            
            results.append({
                'config': config_name,
                'time_window': tw,
                'multiplier': mult,
                'total_outcome': total_outcome,
                'max_mae': max_mae,
                'avg_mae': avg_mae,
                'risk_adjusted_ratio': risk_adjusted_ratio,
                'profit_count': profit_count,
                'loss_count': loss_count,
                'win_rate': win_rate,
                'avg_outcome': total_outcome / total_trades if total_trades > 0 else 0
            })

# Create DataFrame
df_results = pd.DataFrame(results)

print("\n" + "=" * 100)
print("RISK-ADJUSTED RANKINGS (by Risk-Adjusted Ratio)")
print("=" * 100)
df_sorted = df_results.sort_values('risk_adjusted_ratio', ascending=False)
print(df_sorted[['config', 'total_outcome', 'max_mae', 'risk_adjusted_ratio', 'win_rate']].head(15).to_string(index=False))

print("\n" + "=" * 100)
print("HEAT MAP: Risk-Adjusted Ratio (Total Outcome / Max MAE)")
print("=" * 100)

pivot_ratio = df_results.pivot_table(values='risk_adjusted_ratio', index='multiplier', columns='time_window', aggfunc='first')
print("\nMultiplier | 5min | 10min | 15min | 20min | 25min | 30min")
print("-" * 70)
for mult in [0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]:
    row = pivot_ratio.loc[mult]
    print(f"   {mult}x    | {row[5]:4.1f} | {row[10]:4.1f}  | {row[15]:4.1f}  | {row[20]:4.1f}  | {row[25]:4.1f}  | {row[30]:4.1f}")

print("\n" + "=" * 100)
print("COMPARISON: Original vs Risk-Adjusted Rankings")
print("=" * 100)

print("\nTop 10 by TOTAL OUTCOME:")
df_by_outcome = df_results.sort_values('total_outcome', ascending=False).head(10)
for i, (_, row) in enumerate(df_by_outcome.iterrows(), 1):
    print(f"  {i}. {row['config']:12s} | Outcome: {row['total_outcome']:6,d} pts | Max MAE: {row['max_mae']:4d} pts | Ratio: {row['risk_adjusted_ratio']:4.1f}x")

print("\nTop 10 by RISK-ADJUSTED RATIO:")
df_by_ratio = df_results.sort_values('risk_adjusted_ratio', ascending=False).head(10)
for i, (_, row) in enumerate(df_by_ratio.iterrows(), 1):
    print(f"  {i}. {row['config']:12s} | Outcome: {row['total_outcome']:6,d} pts | Max MAE: {row['max_mae']:4d} pts | Ratio: {row['risk_adjusted_ratio']:4.1f}x")

print("\n" + "=" * 100)
print("KEY INSIGHTS & RECOMMENDATIONS")
print("=" * 100)

best_outcome = df_results.loc[df_results['total_outcome'].idxmax()]
best_ratio = df_results.loc[df_results['risk_adjusted_ratio'].idxmax()]

print(f"\n1. ORIGINAL BEST (by Total Outcome):")
print(f"   Config: {best_outcome['config']}")
print(f"   Total Outcome: {best_outcome['total_outcome']:,} pts")
print(f"   Max MAE: {best_outcome['max_mae']} pts")
print(f"   Risk-Adjusted Ratio: {best_outcome['risk_adjusted_ratio']:.2f}x")
print(f"   Win Rate: {best_outcome['win_rate']:.1f}%")

print(f"\n2. NEW BEST (by Risk-Adjusted Ratio):")
print(f"   Config: {best_ratio['config']}")
print(f"   Total Outcome: {best_ratio['total_outcome']:,} pts")
print(f"   Max MAE: {best_ratio['max_mae']} pts")
print(f"   Risk-Adjusted Ratio: {best_ratio['risk_adjusted_ratio']:.2f}x")
print(f"   Win Rate: {best_ratio['win_rate']:.1f}%")

print("\n" + "-" * 70)
if best_outcome['config'] != best_ratio['config']:
    print(f"*** RECOMMENDATION CHANGED! ***")
    print(f"From: {best_outcome['config']} (outcome-focused)")
    print(f"To:   {best_ratio['config']} (risk-adjusted)")
else:
    print(f"*** SAME RECOMMENDATION - Both metrics agree! ***")
    print(f"{best_outcome['config']} is optimal for both total outcome AND risk-adjusted return")
print("-" * 70)

# Create Excel
print("\n" + "=" * 100)
print("Creating Excel file...")
print("=" * 100)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()

# RESULT sheet - sorted by risk-adjusted ratio
ws_result = wb.active
ws_result.title = "RESULT_RiskAdjusted"

ws_result['A1'] = "RISK-ADJUSTED ANALYSIS: Ranked by TotalOutcome / MaxMAE Ratio"
ws_result['A1'].font = Font(size=14, bold=True)
ws_result.merge_cells('A1:K1')
ws_result['A1'].alignment = Alignment(horizontal='center')

# Headers
headers = ['Rank', 'Config', 'TimeWindow', 'Multiplier', 'TotalOutcome', 'MaxMAE', 'RiskAdjRatio', 'WinRate', 'Profit', 'Loss', 'AvgOutcome']
for col, header in enumerate(headers, 1):
    cell = ws_result.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Data
df_export = df_results.sort_values('risk_adjusted_ratio', ascending=False).reset_index(drop=True)
for idx, row in df_export.iterrows():
    ws_result.cell(row=idx+4, column=1, value=idx+1)
    ws_result.cell(row=idx+4, column=2, value=row['config'])
    ws_result.cell(row=idx+4, column=3, value=row['time_window'])
    ws_result.cell(row=idx+4, column=4, value=row['multiplier'])
    ws_result.cell(row=idx+4, column=5, value=row['total_outcome'])
    ws_result.cell(row=idx+4, column=6, value=row['max_mae'])
    ws_result.cell(row=idx+4, column=7, value=round(row['risk_adjusted_ratio'], 2))
    ws_result.cell(row=idx+4, column=8, value=f"{row['win_rate']:.1f}%")
    ws_result.cell(row=idx+4, column=9, value=row['profit_count'])
    ws_result.cell(row=idx+4, column=10, value=row['loss_count'])
    ws_result.cell(row=idx+4, column=11, value=round(row['avg_outcome'], 1))

# Column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    ws_result.column_dimensions[col].width = 15

filename = f"data/RiskAdjusted_Analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
wb.save(filename)
print(f"\nExcel saved: {filename}")

mt5.shutdown()
print("\nDone!")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--date', default='2026-03-20', help='Date to analyze (YYYY-MM-DD)')
    args = parser.parse_args()
    
    from datetime import datetime
    target_date = datetime.strptime(args.date, '%Y-%m-%d').date()
    main(target_date)
